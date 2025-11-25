import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ====================================
# LOAD DATA
# ====================================

script_dir = os.path.dirname(os.path.abspath(__file__))

# Load counterparties CSV (in Entity_classification_results folder)
counterparties_path = os.path.join(script_dir, 'Entity_classification_results', 'counterparties.csv')
df_counterparties = pd.read_csv(counterparties_path)

print(f"Loaded counterparties: {len(df_counterparties)} entities")
print(f"  Path: {counterparties_path}")

# Load EOA addresses JSON (in work directory)
eoa_path = os.path.join(script_dir, 'eoa_addresses.json')
with open(eoa_path, 'r') as f:
    eoa_data = json.load(f)

df_eoa = pd.DataFrame(eoa_data)

print(f"Loaded EOA addresses: {len(df_eoa)} addresses")
print(f"  Path: {eoa_path}")

# ====================================
# CREATE EXCEL FILE WITH TWO SHEETS
# ====================================

output_file = 'Wintermute_Counterparties_Complete.xlsx'

# Create Excel writer
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write Sheet 1: Named Entities & Contracts
    df_counterparties.to_excel(writer, sheet_name='Named Entities', index=False)
    
    # Write Sheet 2: EOA Addresses
    df_eoa.to_excel(writer, sheet_name='EOA Addresses', index=False)

print(f"\nCreated Excel file: {output_file}")

# ====================================
# FORMAT THE EXCEL FILE
# ====================================

# Load the workbook for formatting
wb = load_workbook(output_file)

# Define styles
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_row_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def format_sheet(ws, name):
    """Apply formatting to a worksheet"""
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Alternate row colors
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_row_fill
        
        # Apply borders
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    print(f"  Formatted sheet: {name}")

# Format both sheets
format_sheet(wb['Named Entities'], 'Named Entities')
format_sheet(wb['EOA Addresses'], 'EOA Addresses')

# Add summary info to Named Entities sheet
ws_entities = wb['Named Entities']

# Add summary at the bottom
last_row = ws_entities.max_row + 2
ws_entities[f'A{last_row}'] = 'SUMMARY STATISTICS'
ws_entities[f'A{last_row}'].font = Font(bold=True, size=12)

summary_row = last_row + 1
ws_entities[f'A{summary_row}'] = f'Total Entities: {len(df_counterparties)}'
ws_entities[f'A{summary_row + 1}'] = f'Total Transactions: {df_counterparties["tx_count"].sum():,}'
ws_entities[f'A{summary_row + 2}'] = f'Total Volume: ${df_counterparties["total_usd"].sum():,.2f}'

# Count unique chains
all_chains = set()
for chains_str in df_counterparties["chains"]:
    for chain in str(chains_str).split(","):
        all_chains.add(chain.strip())

ws_entities[f'A{summary_row + 3}'] = f'Unique Chains: {len(all_chains)}'

# Add summary to EOA sheet
ws_eoa = wb['EOA Addresses']

last_row_eoa = ws_eoa.max_row + 2
ws_eoa[f'A{last_row_eoa}'] = 'SUMMARY STATISTICS'
ws_eoa[f'A{last_row_eoa}'].font = Font(bold=True, size=12)

summary_row_eoa = last_row_eoa + 1
ws_eoa[f'A{summary_row_eoa}'] = f'Total EOA Addresses: {len(df_eoa)}'

# Count by status
if 'status' in df_eoa.columns:
    status_counts = df_eoa['status'].value_counts()
    for idx, (status, count) in enumerate(status_counts.items(), start=1):
        ws_eoa[f'A{summary_row_eoa + idx}'] = f'{status}: {count}'

# Save the formatted workbook
wb.save(output_file)

print(f"\n✓ Excel file created and formatted: {output_file}")
print(f"\nFile contains:")
print(f"  Sheet 1: Named Entities ({len(df_counterparties)} rows)")
print(f"  Sheet 2: EOA Addresses ({len(df_eoa)} rows)")

# ====================================
# CREATE SUMMARY STATISTICS
# ====================================

print("\n" + "="*80)
print("COUNTERPARTY BREAKDOWN")
print("="*80)

print(f"\nNamed Entities & Contracts:")
print(f"  Total entities: {len(df_counterparties)}")
print(f"  Total transactions: {df_counterparties['tx_count'].sum():,}")
print(f"  Total volume: ${df_counterparties['total_usd'].sum():,.2f}")

print(f"\nTop 10 by volume:")
for rank, (_, row) in enumerate(df_counterparties.head(10).iterrows(), start=1):
    print(f"  {rank:2d}. {row['entity_name']:30s} ${row['total_usd']:>15,.0f}  ({row['tx_count']:>6,} txs)")


if 'status' in df_eoa.columns:
    print(f"\nEOA Addresses:")
    print(f"  Total addresses: {len(df_eoa)}")
    print(f"\nBy status:")
    for status, count in df_eoa['status'].value_counts().items():
        print(f"  {status}: {count}")

print("\n" + "="*80)